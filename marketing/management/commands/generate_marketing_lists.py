import csv
import os
import re
from datetime import date

from django.core.management.base import BaseCommand
from django.conf import settings

from tenants.models import Empresa
from marketing.models import CarrinhoAbandonado, LeadNaoComprador, Comprador


def format_phone(phone_raw):
    if not phone_raw:
        return ''
    phone = re.sub(r'\D', '', str(phone_raw))
    if len(phone) < 10:
        return ''
    if not phone.startswith('55'):
        phone = f'55{phone}'
    return phone


def clean_postcode(postcode):
    if not postcode:
        return ''
    return re.sub(r'\D', '', str(postcode))


class Command(BaseCommand):
    help = 'Gera listas de marketing (CSV/XLSX) para gestor de trafego'

    def add_arguments(self, parser):
        parser.add_argument(
            '--empresa',
            type=str,
            help='Slug da empresa (ex: tarragona). Se omitido, gera para todas.',
        )
        parser.add_argument(
            '--format',
            type=str,
            default='csv',
            choices=['csv', 'xlsx'],
            help='Formato de saida (default: csv)',
        )
        parser.add_argument(
            '--output-dir',
            type=str,
            default=None,
            help='Diretorio base de saida (default: marketing_exports/)',
        )

    def handle(self, *args, **options):
        empresa_slug = options.get('empresa')
        fmt = options['format']
        base_dir = options.get('output_dir') or os.path.join(settings.BASE_DIR, 'marketing_exports')

        if empresa_slug:
            empresas = Empresa.objects.filter(slug=empresa_slug, ativo=True)
            if not empresas.exists():
                self.stderr.write(f'Empresa "{empresa_slug}" nao encontrada ou inativa.')
                return
        else:
            empresas = Empresa.objects.filter(ativo=True)

        today = date.today().strftime('%Y-%m-%d')

        for empresa in empresas:
            out_dir = os.path.join(base_dir, empresa.slug, today)
            os.makedirs(out_dir, exist_ok=True)

            self.stdout.write(f'\n=== {empresa.nome} ({empresa.slug}) ===')

            self._export_carrinhos(empresa, out_dir, fmt)
            self._export_leads(empresa, out_dir, fmt)
            self._export_compradores(empresa, out_dir, fmt)

            self.stdout.write(self.style.SUCCESS(f'Arquivos gerados em {out_dir}'))

    # ── Carrinhos Abandonados ──────────────────────────────────────

    def _export_carrinhos(self, empresa, out_dir, fmt):
        qs = CarrinhoAbandonado.objects.filter(empresa=empresa)
        count = qs.count()
        self.stdout.write(f'  Carrinhos abandonados: {count}')

        headers = ['phone', 'email', 'fn', 'ln', 'ct', 'st', 'zip', 'country', 'value']

        def row(cart):
            return [
                format_phone(cart.customer.phone),
                cart.customer.email or '',
                cart.customer.first_name or '',
                cart.customer.last_name or '',
                cart.customer.billing_city or '',
                cart.customer.billing_state or '',
                clean_postcode(cart.customer.billing_postcode),
                'BR',
                str(cart.cart_total or ''),
            ]

        path = os.path.join(out_dir, f'carrinhos_abandonados.{fmt}')
        self._write_file(path, fmt, headers, qs, row)

    # ── Leads Nao Compradores ─────────────────────────────────────

    def _export_leads(self, empresa, out_dir, fmt):
        qs = LeadNaoComprador.objects.filter(empresa=empresa)
        count = qs.count()
        self.stdout.write(f'  Leads nao compradores: {count}')

        headers = ['phone', 'fn']

        def row(lead):
            first = (lead.nome or '').split()[0] if lead.nome else ''
            return [format_phone(lead.whatsapp), first]

        path = os.path.join(out_dir, f'leads_nao_compradores.{fmt}')
        self._write_file(path, fmt, headers, qs, row)

    # ── Compradores ───────────────────────────────────────────────

    def _export_compradores(self, empresa, out_dir, fmt):
        qs = Comprador.objects.filter(empresa=empresa)
        count = qs.count()
        self.stdout.write(f'  Compradores: {count}')

        headers = ['phone', 'email', 'fn', 'ln', 'ct', 'st', 'zip', 'country', 'value']

        def row(customer):
            return [
                format_phone(customer.phone),
                customer.email or '',
                customer.first_name or '',
                customer.last_name or '',
                customer.billing_city or '',
                customer.billing_state or '',
                clean_postcode(customer.billing_postcode),
                'BR',
                str(customer.total_spent or ''),
            ]

        path = os.path.join(out_dir, f'compradores.{fmt}')
        self._write_file(path, fmt, headers, qs, row)

    # ── Writer helpers ────────────────────────────────────────────

    def _write_file(self, path, fmt, headers, queryset, row_func):
        if fmt == 'csv':
            self._write_csv(path, headers, queryset, row_func)
        else:
            self._write_xlsx(path, headers, queryset, row_func)

    def _write_csv(self, path, headers, queryset, row_func):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            f.write('\ufeff')  # BOM para Excel
            writer = csv.writer(f)
            writer.writerow(headers)
            for obj in queryset.iterator():
                writer.writerow(row_func(obj))

    def _write_xlsx(self, path, headers, queryset, row_func):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.stderr.write('openpyxl nao instalado. Execute: pip install openpyxl')
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, obj in enumerate(queryset.iterator(), 2):
            for col_idx, val in enumerate(row_func(obj), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(path)
