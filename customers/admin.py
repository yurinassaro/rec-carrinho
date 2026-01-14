from django.contrib import admin
from django.utils import timezone
from django.utils.html import format_html
from django.urls import path
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from import_export import resources
from import_export.admin import ExportMixin
from django.db.models import Count
from .models import Customer, Cart, Order, CustomerAnalysis, Lead
from tenants.admin import TenantAdminMixin
import json
from datetime import datetime
import csv


class CustomerResource(resources.ModelResource):
    class Meta:
        model = Customer
        fields = ('email', 'first_name', 'last_name', 'phone', 'whatsapp_number',
                 'status', 'score', 'total_spent', 'completed_orders', 'abandoned_carts')


@admin.register(Customer)
class CustomerAdmin(TenantAdminMixin, ExportMixin, admin.ModelAdmin):
    resource_class = CustomerResource
    
    list_display = ['email', 'full_name', 'phone_display', 'status_badge', 
                   'score_display', 'total_spent_display', 'last_activity']
    
    list_filter = ['status', 'created_at', 'last_activity']
    
    search_fields = ['email', 'first_name', 'last_name', 'phone']
    
    readonly_fields = ['score', 'status', 'created_at', 'updated_at', 
                      'last_analyzed', 'whatsapp_number']
    
    fieldsets = (
        ('Identificação', {
            'fields': ('email', 'first_name', 'last_name', 'phone', 'whatsapp_number')
        }),
        ('Status e Análise', {
            'fields': ('status', 'score', 'tags', 'notes')
        }),
        ('Estatísticas', {
            'fields': ('total_orders', 'completed_orders', 'total_spent', 
                      'average_order_value', 'abandoned_carts', 'total_abandoned_value')
        }),
        ('Datas', {
            'fields': ('first_seen', 'last_purchase', 'last_activity', 
                      'created_at', 'updated_at', 'last_analyzed')
        }),
    )
    
    def phone_display(self, obj):
        if obj.whatsapp_number:
            return format_html(
                '<a href="https://wa.me/{}" target="_blank">{}</a>',
                obj.whatsapp_number, obj.phone
            )
        return obj.phone
    phone_display.short_description = 'Telefone'
    
    def status_badge(self, obj):
        colors = {
            'never_bought': 'gray',
            'first_time': 'blue',
            'returning': 'green',
            'abandoned_only': 'orange',
            'inactive': 'red',
            'vip': 'gold',
        }
        color = colors.get(obj.status, 'gray')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def score_display(self, obj):
        return format_html(
            '<div style="width:100px; background:#ddd;">'
            '<div style="width:{}%; background:#4CAF50; text-align:center; color:white;">{}</div>'
            '</div>',
            obj.score, obj.score
        )
    score_display.short_description = 'Score'
    
    def total_spent_display(self, obj):
        return f'R$ {obj.total_spent:,.2f}'
    total_spent_display.short_description = 'Total Gasto'
    
    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('carts', 'orders')
    
    actions = ['export_whatsapp_list', 'mark_as_contacted']
    
    def export_whatsapp_list(self, request, queryset):
        """Exporta lista para WhatsApp"""
        customers = queryset.filter(phone__isnull=False).exclude(phone='')
        
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="whatsapp_list.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nome', 'WhatsApp', 'Email', 'Status', 'Score'])
        
        for customer in customers:
            if customer.whatsapp_number:
                writer.writerow([
                    customer.full_name,
                    customer.whatsapp_number,
                    customer.email,
                    customer.get_status_display(),
                    customer.score
                ])
        
        return response
    
    export_whatsapp_list.short_description = "Exportar lista WhatsApp"


@admin.register(CustomerAnalysis)
class CustomerAnalysisAdmin(admin.ModelAdmin):
    list_display = ['date', 'total_customers', 'new_customers', 
                   'abandoned_carts', 'total_revenue_display']
    
    list_filter = ['date']
    
    def total_revenue_display(self, obj):
        return f'R$ {obj.total_revenue:,.2f}'
    total_revenue_display.short_description = 'Receita Total'

@admin.register(Cart)
class CartAdmin(TenantAdminMixin, admin.ModelAdmin):
    list_display = ['get_customer_email', 'get_customer_phone', 'cart_total',
                   'status_dropdown', 'email_toggle', 'whatsapp_toggle', 'created_at']
    list_filter = ['status', 'recovery_email_sent', 'recovery_whatsapp_sent',
                  'was_recovered', 'created_at']
    search_fields = ['customer__email', 'customer__phone', 'checkout_id']

    def get_customer_email(self, obj):
        return obj.customer.email
    get_customer_email.short_description = 'Email'

    def get_customer_phone(self, obj):
        return obj.customer.whatsapp_number or obj.customer.phone or '-'
    get_customer_phone.short_description = 'Telefone'

    def status_dropdown(self, obj):
        """Dropdown para alterar status do carrinho via AJAX"""
        from django.utils.safestring import mark_safe

        colors = {
            'active': '#2196F3',
            'abandoned': '#ff9800',
            'recovered': '#4CAF50',
            'converted': '#9C27B0',
        }

        options_html = []
        for value, label in Cart.CART_STATUS:
            selected = 'selected' if value == obj.status else ''
            options_html.append(f'<option value="{value}" {selected}>{label}</option>')

        select_html = f'''
            <select class="cart-status-select"
                    data-cart-id="{obj.id}"
                    style="
                        background: {colors.get(obj.status, '#666')};
                        color: white;
                        border: none;
                        padding: 5px 10px;
                        border-radius: 4px;
                        font-weight: bold;
                        cursor: pointer;
                        font-size: 12px;
                    ">
                {''.join(options_html)}
            </select>
        '''
        return mark_safe(select_html)

    status_dropdown.short_description = 'Status'
    
    def email_toggle(self, obj):
        """Botão toggle para marcar envio de email"""
        if obj.recovery_email_sent:
            btn_color = '#4CAF50'
            btn_text = '✅ Email Enviado'
            if obj.recovery_email_date:
                date_text = f'<br><small style="opacity: 0.8;">{obj.recovery_email_date.strftime("%d/%m %H:%M")}</small>'
            else:
                date_text = ''
        else:
            btn_color = '#f44336'
            btn_text = '❌ Não Enviado'
            date_text = ''
        
        return format_html(
            '''
            <button onclick="toggleRecovery({}, 'email', {})" 
                    style="background: {}; color: white; border: none; 
                        padding: 8px 12px; border-radius: 4px; cursor: pointer;
                        font-size: 12px; min-width: 140px; text-align: center;">
                <div>{}</div>
                {}
            </button>
            ''',
            obj.id, 
            'false' if obj.recovery_email_sent else 'true',
            btn_color,
            btn_text,
            format_html(date_text) if date_text else ''
        )
    
    email_toggle.short_description = '📧 Email'
    email_toggle.allow_tags = True  # Adicionar esta linha

    def whatsapp_toggle(self, obj):
        """Botão toggle para marcar envio de WhatsApp"""
        if not obj.customer.whatsapp_number:
            return format_html('<span style="color: #999;">Sem WhatsApp</span>')

        if obj.recovery_whatsapp_sent:
            btn_color = '#25D366'
            btn_text = '✅ WhatsApp Enviado'
            if obj.recovery_whatsapp_date:
                date_text = f'<br><small style="opacity: 0.8;">{obj.recovery_whatsapp_date.strftime("%d/%m %H:%M")}</small>'
            else:
                date_text = ''
        else:
            btn_color = '#999'
            btn_text = '❌ Não Enviado'
            date_text = ''

        # Mensagem personalizada com nome do cliente
        primeiro_nome = obj.customer.first_name.split()[0] if obj.customer.first_name else "tudo"

        # Pegar mensagem configurada da empresa
        msg_template = 'Olá {nome}, tudo bem ??'
        if obj.empresa and hasattr(obj.empresa, 'msg_whatsapp_cart'):
            msg_template = obj.empresa.msg_whatsapp_cart or msg_template

        # Escapar caracteres especiais para JavaScript
        msg_escaped = msg_template.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")

        return format_html(
            '''
            <div style="display: flex; align-items: center; gap: 5px;">
                <button onclick="toggleRecovery({}, 'whatsapp', {})"
                        style="background: {}; color: white; border: none;
                            padding: 8px 12px; border-radius: 4px; cursor: pointer;
                            font-size: 12px; min-width: 140px; text-align: center;">
                    <div>{}</div>
                    {}
                </button>
                <button onclick="openWhatsApp('{}', '{}', '{}', '{}')"
                        style="background: #25D366; color: white; border: none;
                            padding: 8px; border-radius: 4px; cursor: pointer;
                            font-size: 16px;" title="Abrir WhatsApp">
                    📱
                </button>
            </div>
            ''',
            obj.id,
            'false' if obj.recovery_whatsapp_sent else 'true',
            btn_color,
            btn_text,
            format_html(date_text) if date_text else '',
            obj.customer.whatsapp_number,
            obj.id,
            primeiro_nome,
            msg_escaped
        )
    whatsapp_toggle.short_description = '📱 WhatsApp'
    whatsapp_toggle.allow_tags = True

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('toggle-recovery/',
                 self.admin_site.admin_view(self.toggle_recovery),
                 name='toggle_recovery'),
            path('update-cart-status/',
                 self.admin_site.admin_view(self.update_cart_status),
                 name='update_cart_status'),
        ]
        return custom_urls + urls
    
    @csrf_exempt
    def toggle_recovery(self, request):
        """Ajax endpoint para toggle de status"""
        if request.method == 'POST':
            data = json.loads(request.body)
            cart_id = data.get('cart_id')
            field_type = data.get('type')  # 'email' ou 'whatsapp'
            new_status = data.get('status')  # true ou false
            
            cart = get_object_or_404(Cart, id=cart_id)
            
            if field_type == 'email':
                cart.recovery_email_sent = new_status
                if new_status:
                    cart.recovery_email_date = timezone.now()
                    cart.recovery_attempts = (cart.recovery_attempts or 0) + 1
                else:
                    cart.recovery_email_date = None
            
            elif field_type == 'whatsapp':
                cart.recovery_whatsapp_sent = new_status
                if new_status:
                    cart.recovery_whatsapp_date = timezone.now()
                    cart.recovery_attempts = (cart.recovery_attempts or 0) + 1
                else:
                    cart.recovery_whatsapp_date = None
            
            cart.save()
            
            return JsonResponse({
                'success': True,
                'new_status': new_status,
                'date': timezone.now().strftime('%d/%m %H:%M')
            })

        return JsonResponse({'success': False})

    @csrf_exempt
    def update_cart_status(self, request):
        """Endpoint AJAX para atualizar status do carrinho"""
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                cart_id = data.get('cart_id')
                new_status = data.get('status')

                cart = get_object_or_404(Cart, id=cart_id)
                cart.status = new_status
                cart.save()

                # Definir cores para cada status
                colors = {
                    'active': '#2196F3',
                    'abandoned': '#ff9800',
                    'recovered': '#4CAF50',
                    'converted': '#9C27B0',
                }

                return JsonResponse({
                    'success': True,
                    'color': colors.get(new_status, '#666'),
                    'status_display': cart.get_status_display()
                })

            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)

        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    class Media:
        js = ('admin/js/cart_recovery.js',)
        css = {
            'all': ('admin/css/cart_admin.css',)
        }

# Adicionar no customers/admin.py

@admin.register(Lead)
class LeadAdmin(TenantAdminMixin, admin.ModelAdmin):
    list_display = ['nome', 'whatsapp_display', 'numero_sapato',
                   'status_dropdown', 'is_customer_badge', 'whatsapp_action', 'created_at']

    list_filter = ['status', 'is_customer', 'numero_sapato', 'created_at']
    search_fields = ['nome', 'whatsapp']
    readonly_fields = ['is_customer', 'related_customer', 'converted_to_customer']
    actions = ['export_leads_action']
    
    def whatsapp_display(self, obj):
        return obj.whatsapp or '-'
    whatsapp_display.short_description = 'WhatsApp'
    
    def numero_sapato(self, obj):
        return obj.numero_sapato or '-'
    numero_sapato.short_description = '👟 Nº'
    
    def status_dropdown(self, obj):
        """Dropdown para alterar status diretamente na lista"""
        from django.utils.safestring import mark_safe

        # Lista de status disponíveis
        status_choices = [
            ('new', 'Novo'),
            ('contacted', 'Contactado'),
            ('customer', 'É Cliente'),
            ('potential', 'Potencial'),
            ('lost', 'Perdido'),
        ]

        # Cores para cada status
        colors = {
            'new': '#2196F3',
            'contacted': '#FF9800',
            'customer': '#4CAF50',
            'potential': '#9C27B0',
            'lost': '#f44336'
        }

        options_html = ''
        for value, label in status_choices:
            selected = 'selected' if value == obj.status else ''
            options_html += f'<option value="{value}" {selected}>{label}</option>'

        html = f'''
            <select class="lead-status-select"
                    data-lead-id="{obj.id}"
                    style="background: {colors.get(obj.status, '#666')}; color: white; border: none;
                           padding: 5px 10px; border-radius: 4px; cursor: pointer;
                           font-weight: bold; font-size: 12px;">
                {options_html}
            </select>
        '''

        return mark_safe(html)
    status_dropdown.short_description = 'Status'
    
    def is_customer_badge(self, obj):
        if obj.is_customer:
            return format_html(
                '<span style="background: #4CAF50; color: white; padding: 3px 10px; '
                'border-radius: 3px;">✅ Cliente</span>'
            )
        return format_html(
            '<span style="background: #FF9800; color: white; padding: 3px 10px; '
            'border-radius: 3px;">🆕 Prospect</span>'
        )
    is_customer_badge.short_description = 'Tipo'
    
    def whatsapp_action(self, obj):
        if not obj.whatsapp_formatted:
            return '-'

        # Mensagem configurada da empresa
        import urllib.parse
        # Pegar apenas o primeiro nome
        primeiro_nome = obj.nome.split()[0] if obj.nome else "tudo"

        # Pegar mensagem configurada da empresa
        msg_template = 'Olá {nome}, tudo bem ??'
        if obj.empresa and hasattr(obj.empresa, 'msg_whatsapp_lead'):
            msg_template = obj.empresa.msg_whatsapp_lead or msg_template

        # Substituir {nome} e tratar quebras de linha
        mensagem = msg_template.replace('{nome}', primeiro_nome)
        msg_encoded = urllib.parse.quote(mensagem)
        whatsapp_url = f"whatsapp://send?phone={obj.whatsapp_formatted}&text={msg_encoded}"

        if obj.whatsapp_sent:
            date_info = obj.whatsapp_sent_date.strftime("%d/%m %H:%M") if obj.whatsapp_sent_date else ''
            return format_html(
                '''
                <div style="display: flex; align-items: center; gap: 5px;">
                    <button onclick="toggleLeadWhatsApp({}, false)"
                            style="background: #25D366; color: white; border: none;
                                   padding: 8px 12px; border-radius: 4px; cursor: pointer;
                                   font-size: 12px; min-width: 140px; text-align: center;">
                        <div>✅ WhatsApp Enviado</div>
                        <small style="opacity: 0.8;">{}</small>
                    </button>
                    <a href="{}"
                       style="background: #25D366; color: white; border: none;
                              padding: 8px; border-radius: 4px; cursor: pointer;
                              font-size: 16px; text-decoration: none; display: inline-block;">
                        📱
                    </a>
                </div>
                ''',
                obj.id,
                date_info,
                whatsapp_url
            )
        else:
            return format_html(
                '''
                <div style="display: flex; align-items: center; gap: 5px;">
                    <button onclick="toggleLeadWhatsApp({}, true)"
                            style="background: #999; color: white; border: none;
                                   padding: 8px 12px; border-radius: 4px; cursor: pointer;
                                   font-size: 12px; min-width: 140px; text-align: center;">
                        ❌ Não Enviado
                    </button>
                    <a href="{}" onclick="setTimeout(() => toggleLeadWhatsApp({}, true), 1000)"
                       style="background: #25D366; color: white; border: none;
                              padding: 8px; border-radius: 4px; cursor: pointer;
                              font-size: 16px; text-decoration: none; display: inline-block;">
                        📱
                    </a>
                </div>
                ''',
                obj.id,
                whatsapp_url,
                obj.id
            )
    whatsapp_action.short_description = 'Ação'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('toggle-lead-whatsapp/',
                 self.admin_site.admin_view(self.toggle_lead_whatsapp),
                 name='toggle_lead_whatsapp'),
            path('update-lead-status/',
                 self.admin_site.admin_view(self.update_lead_status),
                 name='update_lead_status'),
            path('export-leads/',
                 self.admin_site.admin_view(self.export_leads_view),
                 name='export_leads'),
            path('export-leads/download/',
                 self.admin_site.admin_view(self.export_leads_download),
                 name='export_leads_download'),
        ]
        return custom_urls + urls

    @csrf_exempt
    def toggle_lead_whatsapp(self, request):
        """Ajax endpoint para toggle de WhatsApp em leads"""
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                lead_id = data.get('lead_id')
                new_status = data.get('status')

                lead = Lead.objects.get(id=lead_id)
                lead.whatsapp_sent = new_status

                if new_status:
                    lead.whatsapp_sent_date = timezone.now()
                    lead.contact_attempts = (lead.contact_attempts or 0) + 1

                    # Atualizar status para 'contacted' APENAS se estiver 'new'
                    if lead.status == 'new':
                        lead.status = 'contacted'
                else:
                    lead.whatsapp_sent_date = None

                lead.save()

                return JsonResponse({
                    'success': True,
                    'new_status': new_status,
                    'date': timezone.now().strftime('%d/%m %H:%M')
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)

        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    @csrf_exempt
    def update_lead_status(self, request):
        """Ajax endpoint para atualizar status do lead"""
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                lead_id = data.get('lead_id')
                new_status = data.get('status')

                lead = Lead.objects.get(id=lead_id)
                lead.status = new_status

                # Se marcar como cliente, atualizar flag
                if new_status == 'customer':
                    lead.is_customer = True

                lead.save()

                # Retornar cor do novo status
                colors = {
                    'new': '#2196F3',
                    'contacted': '#FF9800',
                    'customer': '#4CAF50',
                    'potential': '#9C27B0',
                    'lost': '#f44336'
                }

                return JsonResponse({
                    'success': True,
                    'new_status': new_status,
                    'color': colors.get(new_status, '#666')
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)

        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    def export_leads_view(self, request):
        """View para escolher formato e filtros de exportação"""
        from django.shortcuts import render
        from datetime import datetime, timedelta

        # Definir datas padrão (últimos 30 dias)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        context = {
            'title': 'Exportar Leads',
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'has_selected': 'export_lead_ids' in request.session,
            'selected_count': len(request.session.get('export_lead_ids', [])),
        }

        return render(request, 'admin/leads_export.html', context)

    def export_leads_download(self, request):
        """Processar download da exportação"""
        from django.http import HttpResponse

        # Obter parâmetros
        export_format = request.GET.get('format', 'csv')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        use_selected = request.GET.get('use_selected') == 'true'

        # Construir queryset
        if use_selected and 'export_lead_ids' in request.session:
            lead_ids = request.session.get('export_lead_ids', [])
            queryset = Lead.objects.filter(id__in=lead_ids)
        else:
            queryset = Lead.objects.all()

            # Aplicar filtros de data
            if start_date:
                queryset = queryset.filter(created_at__gte=start_date)
            if end_date:
                queryset = queryset.filter(created_at__lte=f"{end_date} 23:59:59")

        # Ordenar por data de criação
        queryset = queryset.order_by('-created_at')

        # Exportar baseado no formato
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'txt':
            return self._export_txt(queryset)
        elif export_format == 'xlsx':
            return self._export_xlsx(queryset)
        elif export_format == 'google_sheets':
            return self._export_google_sheets(queryset)

        return HttpResponse('Formato inválido', status=400)

    def _export_csv(self, queryset):
        """Exportar para CSV"""
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # Adicionar BOM para Excel reconhecer UTF-8
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(['Nome', 'WhatsApp', 'Número Sapato', 'Status', 'É Cliente', 'WhatsApp Enviado', 'Data Criação'])

        for lead in queryset:
            writer.writerow([
                lead.nome,
                lead.whatsapp,
                lead.numero_sapato,
                lead.get_status_display(),
                'Sim' if lead.is_customer else 'Não',
                'Sim' if lead.whatsapp_sent else 'Não',
                lead.created_at.strftime('%d/%m/%Y %H:%M') if lead.created_at else ''
            ])

        return response

    def _export_txt(self, queryset):
        """Exportar para TXT"""
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt"'

        # Adicionar BOM para reconhecer UTF-8
        response.write('\ufeff')

        lines = []
        lines.append('=' * 80)
        lines.append(f'EXPORTAÇÃO DE LEADS - {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        lines.append('=' * 80)
        lines.append('')

        for lead in queryset:
            lines.append(f'Nome: {lead.nome}')
            lines.append(f'WhatsApp: {lead.whatsapp}')
            lines.append(f'Número do Sapato: {lead.numero_sapato}')
            lines.append(f'Status: {lead.get_status_display()}')
            lines.append(f'É Cliente: {"Sim" if lead.is_customer else "Não"}')
            lines.append(f'WhatsApp Enviado: {"Sim" if lead.whatsapp_sent else "Não"}')
            lines.append(f'Data de Criação: {lead.created_at.strftime("%d/%m/%Y %H:%M") if lead.created_at else ""}')
            lines.append('-' * 80)
            lines.append('')

        response.write('\n'.join(lines))
        return response

    def _export_xlsx(self, queryset):
        """Exportar para XLSX"""
        from django.http import HttpResponse
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('Biblioteca openpyxl não instalada. Execute: pip install openpyxl', status=500)

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Cabeçalhos
        headers = ['Nome', 'WhatsApp', 'Número Sapato', 'Status', 'É Cliente', 'WhatsApp Enviado', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Dados
        for row_idx, lead in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=lead.nome)
            ws.cell(row=row_idx, column=2, value=lead.whatsapp)
            ws.cell(row=row_idx, column=3, value=lead.numero_sapato)
            ws.cell(row=row_idx, column=4, value=lead.get_status_display())
            ws.cell(row=row_idx, column=5, value='Sim' if lead.is_customer else 'Não')
            ws.cell(row=row_idx, column=6, value='Sim' if lead.whatsapp_sent else 'Não')
            ws.cell(row=row_idx, column=7, value=lead.created_at.strftime('%d/%m/%Y %H:%M') if lead.created_at else '')

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20

        # Salvar
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)

        return response

    def _export_google_sheets(self, queryset):
        """Exportar para Google Sheets - retorna CSV para importar"""
        # Por enquanto, retornar CSV que pode ser importado no Google Sheets
        # Para integração real, seria necessário configurar Google API
        # Retorna o CSV diretamente para que o usuário possa importar no Google Sheets
        return self._export_csv(queryset)

    def export_leads_action(self, request, queryset):
        """Action para exportar leads selecionados"""
        # Armazenar IDs dos leads selecionados na sessão
        request.session['export_lead_ids'] = list(queryset.values_list('id', flat=True))
        # Redirecionar para página de exportação
        from django.shortcuts import redirect
        return redirect('/admin/customers/lead/export-leads/')
    export_leads_action.short_description = "📊 Exportar leads selecionados"

    class Media:
        js = ('admin/js/lead_admin.js',)